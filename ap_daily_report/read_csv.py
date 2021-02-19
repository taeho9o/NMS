from openpyxl import load_workbook


def load_wb(file_name):
    try:
        if not file_name.endswith('.xlsx'):
            file_name = file_name + '.xlsx'
        wb = load_workbook(file_name)
        print('load xlsx file')
        return wb
    except Exception as e:
        print(e)


def save_wb(wb, file_name):
    try:
        if not file_name.endswith('.xlsx'):
            file_name = file_name + '.xlsx'
        wb.save(file_name)
    except Exception as e:
        print(e)


# get sheet name + del result sheet
def get_sheet_names_list(wb):
    try:
        sheet_names_list = wb.get_sheet_names()
        print('get sheet names')
        return sheet_names_list
    except Exception as e:
        print(e)


def read_xlsx(wb, sheet_name):
    try:
        ws = wb[sheet_name]

        host_info_list = []
        i = 1

        while len(host_info_list) < 30:
            hostname = ws.cell(row=i, column=2).value
            max_clnt = ws.cell(row=i, column=3).value
            unq_clnt = ws.cell(row=i, column=4).value
            i += 1
            if not hostname.startswith('W_'):
                continue
            if hostname is None:
                break
            host_info = [hostname, max_clnt, unq_clnt]
            host_info_list.append(host_info)
        return host_info_list
    except Exception as e:
        print(e)


def create_result(wb, list_a):
    try:
        # make 'result' sheet
        wb.create_sheet(title='result')
        # open 'result' sheet
        ws = wb['result']

        # fill 'result' sheet
        for i in range(0, len(list_a)):
            ws[chr(65) + str(i + 1)] = list_a[i]
            ws[chr(66) + str(i + 1)] = 0
        print('create result sheet')
        # wb.save('report.xlsx')
    except Exception as e:
        print(e)


def delete_result_sheet(wb, sheet_name):
    wb.remove(wb[sheet_name])
    print('delete result sheet')


def count_cnt(wb, ap_name):
    try:
        ws = wb['result']
        max_row = ws.max_row + 1
        for i in range(1, max_row):
            if ws.cell(row=i, column=1).value == ap_name:
                ws[chr(66) + str(i)] = ws[chr(66) + str(i)].value + 1

        print('count result')
        # wb.save('report.xlsx')
    except Exception as e:
        print(e)